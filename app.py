"""Data Sheet 비교 분석 AI 에이전트 - Streamlit 웹 UI"""
import sys
import os
import io
import tempfile
from pathlib import Path
from datetime import datetime

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))

from parsers.pdf_parser import parse_pdf
from parsers.excel_parser import parse_excel
from parsers.csv_parser import parse_csv
from parsers.image_parser import parse_image
from agent.analyzer import DataSheetAnalyzer

# ── 언어 텍스트 사전 ─────────────────────────────────────
T = {
    "ko": {
        "page_title": "Data Sheet 비교 분석",
        "main_title": "📊 Data Sheet 비교 분석 AI",
        "sub_title": "PDF · Excel · CSV 파일을 업로드하면 AI가 자동으로 스펙을 비교 분석합니다.",
        "sidebar_upload": "📁 파일 업로드",
        "upload_label": "비교할 Data Sheet 파일을 선택하세요",
        "upload_help": "PDF, Excel, CSV 형식 지원. 최소 2개 이상 업로드하세요.",
        "settings": "⚙️ 설정",
        "show_raw": "상세 JSON 데이터 표시",
        "how_to": "📌 사용 방법",
        "how_to_steps": "1. � 기준품 업로드\n2. 🔍 비교품 업로드\n3. **분석 시작** 버튼 클릭\n4. 결과 확인 및 보고서 다운로드",
        "info_upload": "📌 기준품(현재 사용 부품)을 왼쪽에, 🔍 비교품(교체 후보)을 오른쪽에 업로드해주세요.",
        "warn_min": "⚠️ 비교 분석을 위해 최소 **2개** 이상의 파일을 업로드해주세요.",
        "uploaded_files": "📂 업로드된 파일",
        "btn_analyze": "🚀 AI 분석 시작",
        "parsing": "파일 파싱 중...",
        "parsing_file": "📄 `{}` 파싱 중...",
        "parsing_done": "✅ `{}` 완료",
        "parsing_fail": "❌ `{}` 파싱 실패: {}",
        "parse_complete": "파싱 완료!",
        "analyzing": "🤖 AI가 Data Sheet를 분석 중입니다... (최대 1분 소요)",
        "analyze_err": "❌ 분석 오류: {}",
        "analyze_done": "✅ 분석이 완료되었습니다!",
        "summary": "📝 분석 요약",
        "products": "🏷 제품별 상세 정보",
        "key_specs": "**주요 사양**",
        "performance": "**성능 지표**",
        "highlights": "**핵심 특징**",
        "comparison": "📊 항목별 비교 테이블",
        "compatibility": "🔗 동일 회로 대체 사용 가능 여부 분석",
        "verdict": "대체 판정",
        "substitutable": "동일 회로 대체 가능",
        "risk_level": "교체 위험도",
        "pin_compatible": "핀 호환성",
        "package_compatible": "패키지 호환성",
        "electrical_compatible": "전기적 파라미터 호환성",
        "possible": "✅ 가능",
        "not_possible": "❌ 불가",
        "yes": "✅ 호환",
        "no": "❌ 불일치",
        "circuit_mods": "회로 수정 필요 사항",
        "safe_conditions": "안전한 교체 조건",
        "cautions": "⚠️ 주의사항 (미준수 시 고장 위험)",
        "performance_impact": "📊 성능 영향",
        "recommendation": "💡 종합 의견",
        "json_raw": "🔍 원본 JSON 데이터",
        "download": "📥 보고서 다운로드",
        "dl_excel": "📊 Excel 보고서 다운로드",
        "dl_md": "📄 Markdown 보고서 다운로드",
        "lang_label": "🌐 언어 / Language",
        "file_label": "파일",
        "product_label": "제품",
        "llm_lang": "Korean",
        "item_label": "항목",
        "ref_label": "기준품",
        "ref_hint": "현재 회로에 실제 사용 중인 부품을 업로드하세요 (1개 이상)",
        "manufacturer": "제조사",
        "ref_badge": "📌 기준품",
        "cmp_label": "비교품",
        "cmp_hint": "교체 검토 중인 후보 부품을 업로드하세요 (1개 이상)",
        "cmp_badge": "🔍 비교품",
        "warn_ref": "⚠️ 기준품(현재 사용 부품)을 업로드해주세요.",
        "warn_cmp": "⚠️ 비교품(교체 후보 부품)을 1개 이상 업로드해주세요.",
        "drag_drop_hint": "PDF · Excel · CSV — 파일을 끌어다 놓거나 클릭하여 선택",
        "schematic_label": "회로도 업로드 (선택사항)",
        "schematic_hint": "회로도 이미지를 첨부하면 실제 회로 기반으로 호환성을 더 정밀하게 분석합니다 (PNG · JPG · PDF)",
        "schematic_badge": "🔌 회로도",
        "parsing_image": "🖼️ `{}` 회로도 읽는 중...",
        "parsing_image_done": "✅ `{}` 완료",
        "prev_result_title": "✅ 이전 분석 결과가 있습니다",
        "prev_result_hint": "파일을 다시 업로드하지 않아도 아래에서 결과를 확인하고 다운로드할 수 있습니다.",
        "session_warning": "⚠️ 페이지 새로고침 또는 서버 재시작 시 결과가 사라집니다. 지금 바로 다운로드하세요.",
        "btn_clear": "🗑️ 결과 초기화",
        "download_warning": "⚠️ 서버 재시작 시 결과가 초기화됩니다 — 지금 저장해 두세요.",
        "circuit_title": "회로도 기반 분석",
        "circuit_no_data": "회로도가 업로드되지 않았습니다. 회로도를 업로드하면 실제 회로 수준의 세밀 분석을 제공합니다.",
        "circuit_locations": "회로상 부품 위치",
        "circuit_surrounding": "주변 부품",
        "circuit_constraints": "회로 제약 사항",
        "circuit_layout": "PCB 레이아웃 주의사항",
        "circuit_mods_schematic": "회로도 기반 추가 수정 사항",
        "circuit_assessment": "회로도 종합 평가",
        "compat_verdict": {
            "Drop-in Compatible": "✅ 즉시 대체 가능 (Drop-in)",
            "Conditionally Replaceable": "⚠️ 조건부 대체 가능",
            "Not Replaceable": "❌ 대체 불가",
        },
    },
    "en": {
        "page_title": "Data Sheet Comparison",
        "main_title": "📊 Data Sheet Comparison AI",
        "sub_title": "Upload PDF · Excel · CSV files and AI will automatically compare and analyze specs.",
        "sidebar_upload": "📁 Upload Files",
        "upload_label": "Select Data Sheet files to compare",
        "upload_help": "Supports PDF, Excel, CSV. Please upload at least 2 files.",
        "settings": "⚙️ Settings",
        "show_raw": "Show raw JSON data",
        "how_to": "📌 How to Use",
        "how_to_steps": "1. � Upload reference part\n2. 🔍 Upload comparison part(s)\n3. Click **Start Analysis**\n4. View results & download",
        "info_upload": "📌 Upload the reference part on the left, and 🔍 comparison part(s) on the right.",
        "warn_min": "⚠️ Please upload at least **2** files for comparison.",
        "uploaded_files": "📂 Uploaded Files",
        "btn_analyze": "🚀 Start Analysis",
        "parsing": "Parsing files...",
        "parsing_file": "📄 Parsing `{}`...",
        "parsing_done": "✅ `{}` done",
        "parsing_fail": "❌ `{}` parse error: {}",
        "parse_complete": "Parsing complete!",
        "analyzing": "🤖 AI is analyzing Data Sheets... (up to 1 min)",
        "analyze_err": "❌ Analysis error: {}",
        "analyze_done": "✅ Analysis complete!",
        "summary": "📝 Summary",
        "products": "🏷 Product Details",
        "key_specs": "**Key Specs**",
        "performance": "**Performance**",
        "highlights": "**Highlights**",
        "comparison": "📊 Comparison Table",
        "compatibility": "🔗 Circuit Interchangeability Analysis",
        "verdict": "Substitution Verdict",
        "substitutable": "Drop-in Replaceable",
        "risk_level": "Substitution Risk",
        "pin_compatible": "Pin Compatibility",
        "package_compatible": "Package Compatibility",
        "electrical_compatible": "Electrical Compatibility",
        "possible": "✅ Yes",
        "not_possible": "❌ No",
        "yes": "✅ Compatible",
        "no": "❌ Incompatible",
        "circuit_mods": "Circuit Modifications Required",
        "safe_conditions": "Safe Substitution Conditions",
        "cautions": "⚠️ Cautions (Failure Risk if Ignored)",
        "performance_impact": "📊 Performance Impact",
        "recommendation": "💡 Overall Recommendation",
        "json_raw": "🔍 Raw JSON Data",
        "download": "📥 Download Report",
        "dl_excel": "📊 Download Excel Report",
        "dl_md": "📄 Download Markdown Report",
        "lang_label": "🌐 언어 / Language",
        "file_label": "File",
        "product_label": "Product",
        "llm_lang": "English",
        "item_label": "Item",
        "ref_label": "Reference",
        "ref_hint": "Currently used in the circuit — upload one or more files",
        "manufacturer": "Manufacturer",
        "ref_badge": "📌 Reference",
        "cmp_label": "Comparison",
        "cmp_hint": "Replacement candidates — upload one or more files",
        "cmp_badge": "🔍 Comparison",
        "warn_ref": "⚠️ Please upload the reference (currently used) part.",
        "warn_cmp": "⚠️ Please upload at least one comparison part.",
        "drag_drop_hint": "PDF · Excel · CSV — drag & drop or click to browse",
        "schematic_label": "Upload Circuit Diagram (Optional)",
        "schematic_hint": "Attach schematic images or PDF (PNG · JPG · PDF) for circuit-level compatibility analysis",
        "schematic_badge": "🔌 Schematic",
        "parsing_image": "🖼️ Reading `{}` schematic...",
        "parsing_image_done": "✅ `{}` done",
        "prev_result_title": "✅ Previous analysis result is available",
        "prev_result_hint": "You can review and download the result below without re-uploading files.",
        "session_warning": "⚠️ Results are lost on page refresh or server restart. Download now.",
        "btn_clear": "🗑️ Clear Results",
        "download_warning": "⚠️ Results will be lost on server restart — save them now.",
        "circuit_title": "Schematic-Based Analysis",
        "circuit_no_data": "No circuit diagram was uploaded. Upload a schematic for circuit-level analysis.",
        "circuit_locations": "Component Locations",
        "circuit_surrounding": "Surrounding Components",
        "circuit_constraints": "Circuit Constraints",
        "circuit_layout": "PCB Layout Notes",
        "circuit_mods_schematic": "Schematic-Based Modifications",
        "circuit_assessment": "Overall Circuit Assessment",
        "compat_verdict": {
            "Drop-in Compatible": "✅ Drop-in Compatible",
            "Conditionally Replaceable": "⚠️ Conditionally Replaceable",
            "Not Replaceable": "❌ Not Replaceable",
        },
    },
}

# ── 페이지 설정 ───────────────────────────────────────────
st.set_page_config(page_title="Data Sheet 비교 분석", page_icon="📊", layout="wide")

st.markdown("""
<style>
    .main-title { font-size: 2.2rem; font-weight: 800; color: #1a73e8; }
    .sub-title  { font-size: 1rem; color: #666; margin-bottom: 1.5rem; }
    .summary-box {
        background: #f0f7ff; border-left: 4px solid #1a73e8;
        padding: 1rem 1.2rem; border-radius: 6px; margin-bottom: 1rem;
    }
    .recommend-box {
        background: #f0fff4; border-left: 4px solid #34a853;
        padding: 1rem 1.2rem; border-radius: 6px;
    }
    .compat-box {
        background: #fffbf0; border-left: 4px solid #fbbc04;
        padding: 1rem 1.2rem; border-radius: 6px; margin-top: 0.5rem;
    }
    /* ── 업로드 구역: 컴테이너 배경색 ─────────────── */
    /* stHorizontalBlock 안의 첫 번째 bordered container = 기준품 */
    div[data-testid="stHorizontalBlock"] div[data-testid="stVerticalBlockBorderWrapper"]:nth-of-type(1) {
        background: linear-gradient(160deg, #E3F2FD 0%, #BBDEFB 100%) !important;
        border: 2.5px solid #1565C0 !important;
        border-radius: 14px !important;
        box-shadow: 0 3px 12px rgba(21,101,192,0.18) !important;
    }
    /* 두 번째 bordered container = 비교품 */
    div[data-testid="stHorizontalBlock"] div[data-testid="stVerticalBlockBorderWrapper"]:nth-of-type(2) {
        background: linear-gradient(160deg, #FFF3E0 0%, #FFE0B2 100%) !important;
        border: 2.5px solid #E65100 !important;
        border-radius: 14px !important;
        box-shadow: 0 3px 12px rgba(230,81,0,0.18) !important;
    }
    /* 드롭존 기본 테두리 제거 (컴테이너 배경색을 사용하므로) */
    div[data-testid="stHorizontalBlock"] [data-testid="stFileUploadDropzone"] {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        min-height: 80px !important;
    }
    .badge-ref {
        display: inline-block;
        background: #1565C0; color: white;
        border-radius: 12px; padding: 1px 10px;
        font-size: 0.78rem; font-weight: 700; margin-right: 6px;
    }
    .badge-cmp {
        display: inline-block;
        background: #E65100; color: white;
        border-radius: 12px; padding: 1px 10px;
        font-size: 0.78rem; font-weight: 700; margin-right: 6px;
    }
    /* 메트릭 카드 값 텍스트 줄바꿈 허용 */
    [data-testid="stMetricValue"] {
        white-space: normal !important;
        word-break: keep-all !important;
        overflow: visible !important;
        font-size: 1.1rem !important;
        line-height: 1.4 !important;
    }
    [data-testid="stMetricLabel"] {
        white-space: normal !important;
        word-break: keep-all !important;
        font-size: 0.82rem !important;
    }
</style>
""", unsafe_allow_html=True)

# ── 파일 파싱 ────────────────────────────────────────────
PARSERS = {".pdf": parse_pdf, ".xlsx": parse_excel, ".xls": parse_excel, ".csv": parse_csv}

def parse_uploaded(uploaded_file) -> dict:
    suffix = Path(uploaded_file.name).suffix.lower()
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded_file.getbuffer())
        tmp_path = tmp.name
    try:
        data = PARSERS[suffix](tmp_path)
        data["file_name"] = uploaded_file.name   # 원본 파일명 보존
        return data
    finally:
        os.unlink(tmp_path)


# ── Excel 보고서 ─────────────────────────────────────────
import math as _math

def _auto_height(texts: list, col_widths: list, line_h: int = 15, min_h: int = 20, max_h: int = 150) -> int:
    """셀 텍스트 길이와 열 너비를 기반으로 적절한 행 높이를 계산합니다."""
    max_lines = 1
    for text, width in zip(texts, col_widths):
        if not text or str(text) in ("-", ""):
            continue
        for segment in str(text).split("\n"):
            wrapped = _math.ceil(max(len(segment), 1) / max(width - 3, 8))
            max_lines = max(max_lines, wrapped)
    return max(min_h, min(max_lines * line_h + 4, max_h))


def build_excel_report(result: dict, lang_key: str) -> bytes:
    t = T[lang_key]
    wb = Workbook()

    # ── 색상 팔레트 ───────────────────────────────────────
    C_HDR_BG   = "1565C0"   # 진한 파랑 (시트 제목 행)
    C_SEC_BG   = "BBDEFB"   # 연한 파랑 (섹션 헤더)
    C_SEC_FG   = "0D47A1"   # 섹션 텍스트
    C_ODD      = "F5F9FF"   # 홀수 행
    C_EVEN     = "FFFFFF"   # 짝수 행
    C_KEY_BG   = "E8EAF6"   # 항목명 열
    C_GREEN_BG = "E8F5E9"
    C_YELL_BG  = "FFF9C4"
    C_RED_BG   = "FFEBEE"
    C_TITLE_BG = "0D47A1"   # 최상위 타이틀

    F_HDR   = PatternFill("solid", fgColor=C_HDR_BG)
    F_SEC   = PatternFill("solid", fgColor=C_SEC_BG)
    F_ODD   = PatternFill("solid", fgColor=C_ODD)
    F_EVEN  = PatternFill("solid", fgColor=C_EVEN)
    F_KEY   = PatternFill("solid", fgColor=C_KEY_BG)
    F_GREEN = PatternFill("solid", fgColor=C_GREEN_BG)
    F_YELL  = PatternFill("solid", fgColor=C_YELL_BG)
    F_RED   = PatternFill("solid", fgColor=C_RED_BG)
    F_TITLE = PatternFill("solid", fgColor=C_TITLE_BG)

    thin  = Side(style="thin",   color="BDBDBD")
    thick = Side(style="medium", color="1565C0")
    BORDER     = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    BORDER_BOT = Border(left=thin,  right=thin,  top=thin,  bottom=thick)

    def _title_row(ws, row, ncols, text):
        """최상단 타이틀 행"""
        c = ws.cell(row=row, column=1, value=text)
        c.fill = F_TITLE
        c.font = Font(color="FFFFFF", bold=True, size=14)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        if ncols > 1:
            ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        ws.row_dimensions[row].height = 36

    def _sec_row(ws, row, ncols, text):
        """섹션 구분 행"""
        c = ws.cell(row=row, column=1, value=text)
        c.fill = F_SEC
        c.font = Font(color=C_SEC_FG, bold=True, size=11)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = BORDER_BOT
        if ncols > 1:
            ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        ws.row_dimensions[row].height = 24

    def _col_hdr(ws, row, col, text):
        """열 헤더"""
        c = ws.cell(row=row, column=col, value=text)
        c.fill = F_HDR
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.row_dimensions[row].height = 30

    def _kv(ws, row, col_k, col_v, key, value, row_idx=0, fill_v=None, key_w=30, val_w=70):
        """키-값 쌍 행"""
        ck = ws.cell(row=row, column=col_k, value=key)
        ck.fill = F_KEY
        ck.font = Font(bold=True, size=10)
        ck.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ck.border = BORDER

        cv = ws.cell(row=row, column=col_v, value=str(value) if value is not None else "-")
        cv.fill = fill_v if fill_v else (F_ODD if row_idx % 2 == 0 else F_EVEN)
        cv.font = Font(size=10)
        cv.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        cv.border = BORDER
        ws.row_dimensions[row].height = _auto_height([key, value], [key_w, val_w])

    def _data(ws, row, col, text, row_idx=0, bold=False, fill=None, center=False, col_w=34):
        """일반 데이터 셀 (행 높이는 호출 측에서 _set_row_h로 일괄 처리)"""
        c = ws.cell(row=row, column=col, value=str(text) if text is not None else "-")
        c.fill = fill if fill else (F_ODD if row_idx % 2 == 0 else F_EVEN)
        c.font = Font(bold=bold, size=10)
        c.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="top", wrap_text=True, indent=0 if center else 1
        )
        c.border = BORDER

    def _list_section(ws, title, items, ncols):
        if not items:
            return
        r = ws.max_row + 1
        _sec_row(ws, r, ncols, title)
        for idx, item in enumerate(items):
            r = ws.max_row + 1
            bullet = ws.cell(row=r, column=1, value="▸")
            bullet.fill = F_KEY
            bullet.font = Font(bold=True, size=10, color="1565C0")
            bullet.alignment = Alignment(horizontal="center", vertical="top")
            bullet.border = BORDER

            cv = ws.cell(row=r, column=2, value=str(item))
            cv.fill = F_ODD if idx % 2 == 0 else F_EVEN
            cv.font = Font(size=10)
            cv.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
            cv.border = BORDER
            if ncols > 2:
                ws.merge_cells(f"B{r}:{get_column_letter(ncols)}{r}")
            ws.row_dimensions[r].height = _auto_height([item], [65], line_h=15, min_h=22, max_h=200)

    # ══════════════════════════════════════════════════════
    # Sheet 1: 요약
    # ══════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary" if lang_key == "en" else "요약"
    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 80
    ws1.sheet_view.showGridLines = False

    _title_row(ws1, 1, 2, "Data Sheet Comparison Report" if lang_key == "en" else "Data Sheet 비교 분석 보고서")

    # 생성 일시
    r = 2
    ws1.cell(row=r, column=1, value="Generated" if lang_key == "en" else "생성일시").font = Font(bold=True, size=9, color="757575")
    ws1.cell(row=r, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S")).font  = Font(size=9, color="757575")
    ws1.row_dimensions[r].height = 18

    # 요약 텍스트
    r = 3
    _sec_row(ws1, r, 2, t["summary"].replace("📝 ", ""))
    r = 4
    c = ws1.cell(row=r, column=1, value=result.get("summary", "-"))
    ws1.merge_cells(f"A{r}:B{r}")
    c.fill = F_ODD
    c.font = Font(size=10)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.border = BORDER
    ws1.row_dimensions[r].height = 80

    # 제품 목록
    r = 5
    _sec_row(ws1, r, 2, ("Products" if lang_key == "en" else "분석 대상 제품") + f"  ({len(result.get('products', []))})")
    r = 6
    _col_hdr(ws1, r, 1, t["file_label"])
    _col_hdr(ws1, r, 2, t["product_label"])
    for idx, prod in enumerate(result.get("products", []), start=1):
        r = ws1.max_row + 1
        fn = prod.get("file_name", "-")
        pn = prod.get("product_name", "-")
        _data(ws1, r, 1, fn, row_idx=idx, bold=True)
        _data(ws1, r, 2, pn, row_idx=idx)
        ws1.row_dimensions[r].height = _auto_height([fn, pn], [24, 80])

    # 종합 의견
    r = ws1.max_row + 2
    _sec_row(ws1, r, 2, t["recommendation"].replace("💡 ", ""))
    r = ws1.max_row + 1
    c = ws1.cell(row=r, column=1, value=result.get("recommendation", "-"))
    ws1.merge_cells(f"A{r}:B{r}")
    c.fill = PatternFill("solid", fgColor="E8F5E9")
    c.font = Font(size=10)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.border = Border(left=Side(style="medium", color="43A047"),
                      right=thin, top=thin, bottom=thin)
    ws1.row_dimensions[r].height = 100

    # ══════════════════════════════════════════════════════
    # Sheet 2: 제품 상세
    # ══════════════════════════════════════════════════════
    products = result.get("products", [])
    ws2 = wb.create_sheet("Product Details" if lang_key == "en" else "제품 상세")
    ws2.column_dimensions["A"].width = 28
    for i in range(len(products)):
        ws2.column_dimensions[get_column_letter(i + 2)].width = 32
    ws2.sheet_view.showGridLines = False

    ncols2 = len(products) + 1
    _title_row(ws2, 1, ncols2, "Product Details" if lang_key == "en" else "제품 상세 정보")

    # 열 헤더 (제품명)
    r = 2
    _col_hdr(ws2, r, 1, t["item_label"])
    for i, prod in enumerate(products):
        _col_hdr(ws2, r, i + 2, prod.get("product_name", prod.get("file_name", f"Product {i+1}")))
    ws2.freeze_panes = "A3"

    # 제조사 행
    r = ws2.max_row + 1
    _sec_row(ws2, r, ncols2, t["manufacturer"])
    r = ws2.max_row + 1
    _data(ws2, r, 1, t["manufacturer"], row_idx=0, bold=True, fill=F_KEY)
    for i, prod in enumerate(products):
        _data(ws2, r, i + 2, prod.get("manufacturer", "-"), row_idx=0)
    ws2.row_dimensions[r].height = 22

    for section_key, section_label in [
        ("key_specs",   t["key_specs"].replace("**", "")),
        ("performance", t["performance"].replace("**", "")),
    ]:
        r = ws2.max_row + 1
        _sec_row(ws2, r, ncols2, section_label)
        all_keys = []
        for prod in products:
            for k in prod.get(section_key, {}):
                if k not in all_keys:
                    all_keys.append(k)
        for idx, k in enumerate(all_keys):
            r = ws2.max_row + 1
            vals = [prod.get(section_key, {}).get(k, "-") for prod in products]
            _data(ws2, r, 1, k, row_idx=idx, bold=True, fill=F_KEY)
            for i, prod in enumerate(products):
                _data(ws2, r, i + 2, vals[i], row_idx=idx)
            ws2.row_dimensions[r].height = _auto_height([k] + vals, [28] + [32]*len(products))

    r = ws2.max_row + 1
    _sec_row(ws2, r, ncols2, t["highlights"].replace("**", ""))
    max_hl = max((len(p.get("highlights", [])) for p in products), default=0)
    for j in range(max_hl):
        r = ws2.max_row + 1
        label = f"{'특징' if lang_key == 'ko' else 'Feature'} {j+1}"
        hl_vals = [prod.get("highlights", [])[j] if j < len(prod.get("highlights", [])) else "-" for prod in products]
        _data(ws2, r, 1, label, row_idx=j, bold=True, fill=F_KEY)
        for i, prod in enumerate(products):
            _data(ws2, r, i + 2, hl_vals[i], row_idx=j)
        ws2.row_dimensions[r].height = _auto_height([label] + hl_vals, [28] + [32]*len(products))

    # ══════════════════════════════════════════════════════
    # Sheet 3: 비교 테이블
    # ══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Comparison" if lang_key == "en" else "비교 테이블")
    ws3.sheet_view.showGridLines = False
    table = result.get("comparison_table", [])
    if table:
        file_names = list(table[0]["values"].keys())
        ncols3 = len(file_names) + 1
        ws3.column_dimensions["A"].width = 38
        for i in range(len(file_names)):
            ws3.column_dimensions[get_column_letter(i + 2)].width = 34

        _title_row(ws3, 1, ncols3, "Comparison Table" if lang_key == "en" else "항목별 비교 테이블")
        r = 2
        _col_hdr(ws3, r, 1, t["item_label"])
        for i, fn in enumerate(file_names):
            _col_hdr(ws3, r, i + 2, fn)
        ws3.freeze_panes = "A3"
        ws3.auto_filter.ref = f"A2:{get_column_letter(ncols3)}2"

        for idx, row_data in enumerate(table):
            r = ws3.max_row + 1
            cat = row_data["category"]
            vals = [row_data["values"].get(fn, "-") for fn in file_names]
            _data(ws3, r, 1, cat, row_idx=idx, bold=True, fill=F_KEY)
            for i, fn in enumerate(file_names):
                _data(ws3, r, i + 2, vals[i], row_idx=idx)
            ws3.row_dimensions[r].height = _auto_height([cat] + vals, [38] + [34]*len(file_names))

    # ══════════════════════════════════════════════════════
    # Sheet 4: 회로 대체 분석
    # ══════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Interchangeability" if lang_key == "en" else "회로 대체 분석")
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 70
    ws4.sheet_view.showGridLines = False
    ix = result.get("interchangeability", {})

    _title_row(ws4, 1, 2, t["compatibility"].replace("🔗 ", ""))

    verdict_raw    = ix.get("verdict", "")
    verdict_fill   = {"Drop-in Compatible": F_GREEN, "Conditionally Replaceable": F_YELL, "Not Replaceable": F_RED}.get(verdict_raw, None)
    verdict_display = t["compat_verdict"].get(verdict_raw, verdict_raw)
    risk            = ix.get("risk_level", "-")
    risk_fill       = {"None": F_GREEN, "Low": F_GREEN, "Medium": F_YELL, "High": F_RED}.get(risk, None)

    # 판정 대형 셀
    r = 2
    _sec_row(ws4, r, 2, "판정 요약" if lang_key == "ko" else "Verdict Summary")
    r = 3
    c_verdict = ws4.cell(row=r, column=1, value=verdict_display)
    ws4.merge_cells(f"A{r}:B{r}")
    c_verdict.fill = verdict_fill if verdict_fill else F_YELL
    c_verdict.font = Font(bold=True, size=16)
    c_verdict.alignment = Alignment(horizontal="center", vertical="center")
    c_verdict.border = Border(left=Side(style="medium", color="1565C0"),
                               right=Side(style="medium", color="1565C0"),
                               top=Side(style="medium",  color="1565C0"),
                               bottom=Side(style="medium", color="1565C0"))
    ws4.row_dimensions[r].height = 42

    # 핵심 지표 표
    r = ws4.max_row + 1
    _sec_row(ws4, r, 2, "핵심 지표" if lang_key == "ko" else "Key Metrics")
    metrics = [
        (t["risk_level"],         risk,                                                            risk_fill),
        (t["substitutable"],      ("O" if ix.get("substitutable") else "X") + "  " +
                                  (t["possible"] if ix.get("substitutable") else t["not_possible"]), None),
        (t["pin_compatible"],     ("O" if ix.get("pin_compatible") else "X") + "  " +
                                  (t["yes"] if ix.get("pin_compatible") else t["no"]),              None),
        (t["package_compatible"], ("O" if ix.get("package_compatible") else "X") + "  " +
                                  (t["yes"] if ix.get("package_compatible") else t["no"]),         None),
        (t["electrical_compatible"], ("O" if ix.get("electrical_compatible") else "X") + "  " +
                                     (t["yes"] if ix.get("electrical_compatible") else t["no"]),   None),
    ]
    for idx, (label, value, mfill) in enumerate(metrics):
        r = ws4.max_row + 1
        fill_v = mfill if mfill else (F_GREEN if value.startswith("O") else (F_RED if value.startswith("X") else None))
        _kv(ws4, r, 1, 2, label, value, row_idx=idx, fill_v=fill_v)

    # 목록 섹션들
    _list_section(ws4, t["circuit_mods"],
                  ix.get("circuit_modifications_required", []), 2)
    _list_section(ws4, t["safe_conditions"],
                  ix.get("safe_conditions", []), 2)
    _list_section(ws4, t["cautions"].replace("⚠️ ", ""),
                  ix.get("cautions", []), 2)
    _list_section(ws4, t["performance_impact"].replace("📊 ", ""),
                  ix.get("performance_impact", []), 2)

    # ══════════════════════════════════════════════════════
    # Sheet 5: 회로도 기반 분석
    # ══════════════════════════════════════════════════════
    ca = result.get("circuit_analysis", {})
    if ca and ca.get("provided"):
        ws5 = wb.create_sheet("Schematic Analysis" if lang_key == "en" else "회로도 분석")
        ws5.column_dimensions["A"].width = 28
        ws5.column_dimensions["B"].width = 80
        ws5.sheet_view.showGridLines = False
        _title_row(ws5, 1, 2, t["circuit_title"])
        r = 2
        _sec_row(ws5, r, 2, t["circuit_assessment"])
        r = 3
        c_ca = ws5.cell(row=r, column=1, value=ca.get("overall_circuit_assessment", "-"))
        ws5.merge_cells(f"A{r}:B{r}")
        c_ca.fill = PatternFill("solid", fgColor="E8F5E9")
        c_ca.font = Font(size=10)
        c_ca.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        c_ca.border = Border(left=Side(style="medium", color="2E7D32"), right=thin, top=thin, bottom=thin)
        ws5.row_dimensions[r].height = 80
        for key, label in [
            ("component_locations",      t["circuit_locations"]),
            ("surrounding_components",   t["circuit_surrounding"]),
            ("circuit_constraints",      t["circuit_constraints"]),
            ("layout_notes",             t["circuit_layout"]),
            ("schematic_based_modifications", t["circuit_mods_schematic"]),
        ]:
            _list_section(ws5, label, ca.get(key, []), 2)
        ws5.sheet_properties.tabColor = "2E7D32"

    # ── 시트 탭 색상 ─────────────────────────────────────
    ws1.sheet_properties.tabColor = "1565C0"
    ws2.sheet_properties.tabColor = "0288D1"
    ws3.sheet_properties.tabColor = "00838F"
    ws4.sheet_properties.tabColor = (
        "43A047" if verdict_raw == "Drop-in Compatible"
        else "F9A825" if verdict_raw == "Conditionally Replaceable"
        else "E53935"
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Markdown 보고서 ───────────────────────────────────────
def build_markdown_report(result: dict, lang_key: str) -> str:
    t = T[lang_key]
    lines = [f"# {t['main_title']}\n{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"]
    lines.append(f"## {t['summary'].replace('📝 ','')}\n{result.get('summary','-')}\n")
    for prod in result.get("products", []):
        lines.append(f"## {prod.get('product_name', prod.get('file_name','-'))}\n**{t['manufacturer']}**: {prod.get('manufacturer', '-')}  |  **{t['file_label']}**: {prod.get('file_name','-')}\n")
        if prod.get("key_specs"):
            lines.append(f"### {t['key_specs'].replace('**','')}\n" + "\n".join(f"- **{k}**: {v}" for k,v in prod["key_specs"].items()) + "\n")
        if prod.get("performance"):
            lines.append(f"### {t['performance'].replace('**','')}\n" + "\n".join(f"- **{k}**: {v}" for k,v in prod["performance"].items()) + "\n")
        if prod.get("highlights"):
            lines.append(f"### {t['highlights'].replace('**','')}\n" + "\n".join(f"- {h}" for h in prod["highlights"]) + "\n")
    table = result.get("comparison_table", [])
    if table:
        fns = list(table[0]["values"].keys())
        lines.append(f"## {t['comparison'].replace('📊 ','')}\n| {t['item_label']} | " + " | ".join(fns) + " |\n")
        lines.append("|---" * (len(fns)+1) + "|\n")
        for row in table:
            lines.append(f"| {row['category']} | " + " | ".join(str(row["values"].get(fn,"-")) for fn in fns) + " |\n")
    ix = result.get("interchangeability", {})
    if ix:
        verdict_raw = ix.get("verdict", "")
        lines.append(f"\n## {t['compatibility'].replace('🔗 ','')}\n")
        lines.append(f"- **{t['verdict']}**: {t['compat_verdict'].get(verdict_raw, verdict_raw)}\n")
        lines.append(f"- **{t['risk_level']}**: {ix.get('risk_level','-')}\n")
        lines.append(f"- **{t['substitutable']}**: {t['possible'] if ix.get('substitutable') else t['not_possible']}\n")
        lines.append(f"- **{t['pin_compatible']}**: {t['yes'] if ix.get('pin_compatible') else t['no']}\n")
        lines.append(f"- **{t['package_compatible']}**: {t['yes'] if ix.get('package_compatible') else t['no']}\n")
        lines.append(f"- **{t['electrical_compatible']}**: {t['yes'] if ix.get('electrical_compatible') else t['no']}\n")
        for key, label in [
            ("circuit_modifications_required", t["circuit_mods"]),
            ("safe_conditions",                t["safe_conditions"]),
            ("cautions",                       t["cautions"]),
            ("performance_impact",             t["performance_impact"]),
        ]:
            items = ix.get(key, [])
            if items:
                lines.append(f"\n### {label.replace('⚠️ ','').replace('📊 ','')}\n" + "\n".join(f"- {i}" for i in items) + "\n")
    # 회로도 기반 분석 섹션
    ca = result.get("circuit_analysis", {})
    if ca and ca.get("provided"):
        lines.append(f"\n## 🔌 {t['circuit_title']}\n")
        if ca.get("overall_circuit_assessment"):
            lines.append(f"{ca['overall_circuit_assessment']}\n")
        for key, label in [
            ("component_locations",      t["circuit_locations"]),
            ("surrounding_components",   t["circuit_surrounding"]),
            ("circuit_constraints",      t["circuit_constraints"]),
            ("layout_notes",             t["circuit_layout"]),
            ("schematic_based_modifications", t["circuit_mods_schematic"]),
        ]:
            items = ca.get(key, [])
            if items:
                lines.append(f"\n### {label}\n" + "\n".join(f"- {i}" for i in items) + "\n")
    if result.get("recommendation"):
        lines.append(f"\n## {t['recommendation'].replace('💡 ','')}\n{result['recommendation']}\n")
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════
#  사이드바
# ═══════════════════════════════════════════════════════════
with st.sidebar:
    lang_choice = st.radio(T["ko"]["lang_label"], options=["한국어", "English"], horizontal=True, index=0)
    lang_key = "ko" if lang_choice == "한국어" else "en"
    t = T[lang_key]

    st.divider()
    st.caption(t["settings"])
    show_raw = st.toggle(t["show_raw"], value=False)
    st.divider()
    st.caption(t["how_to"])
    st.markdown(t["how_to_steps"])

# ═══════════════════════════════════════════════════════════
#  Session State 초기화
# ═══════════════════════════════════════════════════════════
if "result" not in st.session_state:
    st.session_state.result = None
if "result_lang" not in st.session_state:
    st.session_state.result_lang = "ko"
if "ref_filenames" not in st.session_state:
    st.session_state.ref_filenames = []
if "circuit_image_names" not in st.session_state:
    st.session_state.circuit_image_names = []

# ═══════════════════════════════════════════════════════════
#  메인
# ═══════════════════════════════════════════════════════════
st.markdown(f'<div class="main-title">{t["main_title"]}</div>', unsafe_allow_html=True)
st.markdown(f'<div class="sub-title">{t["sub_title"]}</div>', unsafe_allow_html=True)
st.divider()

# ── 파일 업로드 (2-구역) ──────────────────────────────────
col_ref, col_cmp = st.columns([1, 2])

with col_ref:
    with st.container(border=True):
        st.markdown(
            f'<div style="color:#0D47A1;font-weight:800;font-size:1.08rem;'
            f'margin-bottom:6px">'
            f'📌 {t["ref_label"]}</div>'
            f'<div style="color:#555;font-size:0.82rem;margin-bottom:4px">{t["ref_hint"]}</div>',
            unsafe_allow_html=True,
        )
        ref_files = st.file_uploader(
            "upload_ref", type=["pdf", "xlsx", "xls", "csv"],
            accept_multiple_files=True, key="upload_ref", label_visibility="collapsed",
        )

with col_cmp:
    with st.container(border=True):
        st.markdown(
            f'<div style="color:#BF360C;font-weight:800;font-size:1.08rem;'
            f'margin-bottom:6px">'
            f'🔍 {t["cmp_label"]}</div>'
            f'<div style="color:#555;font-size:0.82rem;margin-bottom:4px">{t["cmp_hint"]}</div>',
            unsafe_allow_html=True,
        )
        cmp_files = st.file_uploader(
            "upload_cmp", type=["pdf", "xlsx", "xls", "csv"],
            accept_multiple_files=True, key="upload_cmp", label_visibility="collapsed",
        )

# ── 회로도 업로드 (선택사항) ──────────────────────────────────
with st.container(border=True):
    st.markdown(
        f'<div style="color:#1B5E20;font-weight:800;font-size:1.0rem;margin-bottom:4px">'
        f'🔌 {t["schematic_label"]}</div>'
        f'<div style="color:#555;font-size:0.82rem;margin-bottom:4px">{t["schematic_hint"]}</div>',
        unsafe_allow_html=True,
    )
    schematic_files = st.file_uploader(
        "upload_schematic", type=["png", "jpg", "jpeg", "webp", "pdf"],
        accept_multiple_files=True, key="upload_schematic", label_visibility="collapsed",
    )
uploaded_files = (ref_files or []) + (cmp_files or [])

if not uploaded_files:
    if st.session_state.result:
        # 파일 없지만 이전 결과 존재 → 프론티로 표시
        col_banner, col_clear = st.columns([4, 1])
        with col_banner:
            st.markdown(
                f'<div style="background:#E8F5E9;border-left:5px solid #43A047;'
                f'padding:12px 16px;border-radius:8px;">'
                f'<strong>{t["prev_result_title"]}</strong><br>'
                f'<span style="font-size:0.88rem;color:#555">{t["prev_result_hint"]}</span><br>'
                f'<span style="font-size:0.82rem;color:#E65100">{t["session_warning"]}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
        with col_clear:
            if st.button(t["btn_clear"], use_container_width=True):
                st.session_state.result = None
                st.session_state.ref_filenames = []
                st.session_state.circuit_image_names = []
                st.rerun()
        st.info(t["info_upload"])
    else:
        st.info(t["info_upload"])
        st.stop()
elif not ref_files:
    st.warning(t["warn_ref"]); st.stop()
elif not cmp_files:
    st.warning(t["warn_cmp"]); st.stop()
else:
    # 업로드된 파일 미리보기
    st.divider()
    total = len(ref_files) + len(cmp_files)
    st.subheader(f"{t['uploaded_files']} ({total})")
    preview_cols = st.columns(total)
    for i, (f, role_badge, badge_class) in enumerate(
        [(f, t["ref_badge"], "badge-ref") for f in ref_files] +
        [(f, t["cmp_badge"], "badge-cmp") for f in cmp_files]
    ):
        with preview_cols[i]:
            with st.container(border=True):
                ext = Path(f.name).suffix.upper()[1:]
                icon = {"PDF": "🔴", "XLSX": "🟢", "XLS": "🟢", "CSV": "🔵"}.get(ext, "⚪")
                st.markdown(
                    f'<span class="{badge_class}">{role_badge}</span>'
                    f'<strong>{icon} {f.name}</strong>',
                    unsafe_allow_html=True,
                )
                st.caption(f"{ext} · {f.size / 1024:.1f} KB")

    # 회로도 미리보기
    if schematic_files:
        st.markdown(f"**🔌 {t['schematic_badge']}** ({len(schematic_files)})")
        sch_cols = st.columns(min(len(schematic_files), 4))
        for i, sf in enumerate(schematic_files):
            with sch_cols[i % 4]:
                if Path(sf.name).suffix.lower() == ".pdf":
                    with st.container(border=True):
                        st.markdown(f"🔴 **{sf.name}**")
                        st.caption(f"PDF · {sf.size / 1024:.1f} KB")
                else:
                    st.image(sf, caption=sf.name, use_container_width=True)
    st.divider()

    if st.button(t["btn_analyze"], type="primary", use_container_width=True):
        parsed_files = []
        roles = []
        circuit_images = []
        with st.status(t["parsing"], expanded=True) as status:
            for f, role in [(f, "reference") for f in ref_files] + [(f, "comparison") for f in cmp_files]:
                st.write(t["parsing_file"].format(f.name))
                try:
                    data = parse_uploaded(f)
                    parsed_files.append(data)
                    roles.append(role)
                    st.write(t["parsing_done"].format(f.name))
                except Exception as e:
                    st.error(t["parsing_fail"].format(f.name, e)); st.stop()
            # 회로도 이미지 파싱
            for f in (schematic_files or []):
                st.write(t["parsing_image"].format(f.name))
                suffix = Path(f.name).suffix.lower()
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    tmp.write(f.getbuffer())
                    tmp_path = tmp.name
                try:
                    img_data = parse_image(tmp_path)
                    img_data["file_name"] = f.name
                    circuit_images.append(img_data)
                    st.write(t["parsing_image_done"].format(f.name))
                except Exception as e:
                    st.error(t["parsing_fail"].format(f.name, e)); st.stop()
                finally:
                    os.unlink(tmp_path)
            status.update(label=t["parse_complete"], state="complete")

        with st.spinner(t["analyzing"]):
            try:
                analyzer = DataSheetAnalyzer()
                result = analyzer.analyze(
                    parsed_files, lang=t["llm_lang"], roles=roles,
                    circuit_images=circuit_images if circuit_images else None,
                )
                st.session_state.result = result
                st.session_state.result_lang = lang_key
                st.session_state.ref_filenames = [f.name for f in ref_files]
                st.session_state.circuit_image_names = [f.name for f in (schematic_files or [])]
            except Exception as e:
                st.error(t["analyze_err"].format(e)); st.stop()

        st.success(t["analyze_done"])
        st.balloons()

# ── 결과 표시 (session_state에서 렌더링) ──────────────────
def render_result(result: dict, t: dict, lang_key: str, show_raw: bool):
    if not result:
        return
    # raw_response 가 있더라도 유효한 JSON 이면 재파싱 시도
    if "raw_response" in result:
        import json as _json, re as _re
        raw = result["raw_response"]
        parsed = None
        for attempt_text in [raw, _re.sub(r'^```(?:json)?\s*|\s*```$', '', raw.strip(), flags=_re.S)]:
            try:
                parsed = _json.loads(attempt_text)
                break
            except Exception:
                m = _re.search(r'\{[\s\S]*\}', attempt_text)
                if m:
                    try:
                        parsed = _json.loads(m.group()); break
                    except Exception:
                        pass
        if isinstance(parsed, dict) and "raw_response" not in parsed:
            result = parsed
            st.session_state.result = result
        else:
            # ── 진단 정보 표시 ──────────────────────────────────
            raw_clean = raw.strip('\ufeff').strip()
            try:
                _json.loads(raw_clean)
            except _json.JSONDecodeError as _e:
                ctx_s = max(0, _e.pos - 80)
                ctx_e = min(len(raw_clean), _e.pos + 80)
                st.error(
                    f"🔍 JSON 파싱 실패 — "
                    f"위치 {_e.pos}/{len(raw_clean)}자, 오류: **{_e.msg}**"
                )
                st.code(
                    raw_clean[ctx_s:_e.pos] + " >>>HERE<<< " + raw_clean[_e.pos:ctx_e],
                    language="text"
                )
            except Exception as _e2:
                st.error(f"기타 오류: {_e2}")
            st.text(raw_clean[:3000])
            return

    ref_names = st.session_state.get("ref_filenames", [])
    st.divider()

    # 요약
    st.subheader(t["summary"])
    st.markdown(f'<div class="summary-box">{result.get("summary","-")}</div>', unsafe_allow_html=True)

    # 제품 카드
    st.subheader(t["products"])
    products = result.get("products", [])
    prod_cols = st.columns(max(len(products), 1))
    for i, (prod, col) in enumerate(zip(products, prod_cols)):
        with col:
            fname = prod.get("file_name", "")
            is_ref = (fname in ref_names) or (i == 0 and not ref_names)
            badge_class = "badge-ref" if is_ref else "badge-cmp"
            badge_text  = t["ref_badge"] if is_ref else t["cmp_badge"]
            with st.container(border=True):
                # ── 역할 구분 컬러 배너 ──────────────────────
                if is_ref:
                    st.markdown(
                        '<div style="background:linear-gradient(90deg,#1565C0,#1976D2);'
                        'color:white;padding:7px 12px;border-radius:6px;'
                        'font-weight:800;font-size:0.88rem;margin-bottom:8px">'
                        f'📌 {t["ref_label"]} — '
                        f'{"현재 사용 중인 부품" if lang_key=="ko" else "Currently in use"}'
                        '</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        '<div style="background:linear-gradient(90deg,#E65100,#F57C00);'
                        'color:white;padding:7px 12px;border-radius:6px;'
                        'font-weight:800;font-size:0.88rem;margin-bottom:8px">'
                        f'🔍 {t["cmp_label"]} — '
                        f'{"교체 후보 부품" if lang_key=="ko" else "Replacement candidate"}'
                        '</div>',
                        unsafe_allow_html=True,
                    )
                st.markdown(
                    f'<strong>📄 {prod.get("product_name", fname)}</strong>',
                    unsafe_allow_html=True,
                )
                mfr = prod.get("manufacturer", "")
                st.caption(f"🏭 {t['manufacturer']}: {mfr}  |  {t['file_label']}: {fname}")
                if prod.get("key_specs"):
                    st.markdown(t["key_specs"])
                    for k, v in prod["key_specs"].items(): st.markdown(f"- `{k}`: {v}")
                if prod.get("performance"):
                    st.markdown(t["performance"])
                    for k, v in prod["performance"].items(): st.markdown(f"- `{k}`: {v}")
                if prod.get("highlights"):
                    st.markdown(t["highlights"])
                    for h in prod["highlights"]: st.markdown(f"- ✅ {h}")

    # 비교 테이블
    table = result.get("comparison_table", [])
    if table:
        st.subheader(t["comparison"])
        fns = list(table[0]["values"].keys())
        # 컬럼명에 📌 기준품 / 🔍 비교품 역할 레이블 추가
        def _col_label(fn):
            if fn in ref_names:
                return f"📌 {fn} ({t['ref_label']})"
            return f"🔍 {fn} ({t['cmp_label']})"
        col_map = {fn: _col_label(fn) for fn in fns}
        rows = [{t["item_label"]: row["category"], **{col_map[fn]: row["values"].get(fn,"-") for fn in fns}} for row in table]
        df = pd.DataFrame(rows).set_index(t["item_label"])
        st.dataframe(df, use_container_width=True, height=min(40*len(df)+50, 600))

    # 🔗 회로 대체 가능 여부 분석
    ix = result.get("interchangeability", {})
    st.subheader(t["compatibility"])
    if ix:
        verdict_raw = ix.get("verdict", "")
        verdict_display = t["compat_verdict"].get(verdict_raw, verdict_raw)
        risk = ix.get("risk_level", "-")

        m1, m2, m3, m4, m5 = st.columns([2, 1.2, 1.2, 1.2, 1.2])
        m1.metric(t["verdict"],            verdict_display)
        m2.metric(t["risk_level"],         risk)
        m3.metric(t["substitutable"],      t["possible"] if ix.get("substitutable")        else t["not_possible"])
        m4.metric(t["pin_compatible"],     t["yes"]      if ix.get("pin_compatible")        else t["no"])
        m5.metric(t["package_compatible"], t["yes"]      if ix.get("package_compatible")    else t["no"])

        st.markdown('<div class="compat-box">', unsafe_allow_html=True)
        elec_ok = ix.get("electrical_compatible")
        st.markdown(f"**{t['electrical_compatible']}**: {t['yes'] if elec_ok else t['no']}")
        if ix.get("circuit_modifications_required"):
            st.markdown(f"**{t['circuit_mods']}**")
            for c in ix["circuit_modifications_required"]: st.markdown(f"- 🔧 {c}")
        if ix.get("safe_conditions"):
            st.markdown(f"**{t['safe_conditions']}**")
            for c in ix["safe_conditions"]: st.markdown(f"- ✅ {c}")
        if ix.get("cautions"):
            st.markdown(f"**{t['cautions']}**")
            for c in ix["cautions"]: st.markdown(f"- ⚠️ {c}")
        if ix.get("performance_impact"):
            st.markdown(f"**{t['performance_impact']}**")
            for c in ix["performance_impact"]: st.markdown(f"- 📊 {c}")
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.info("ℹ️ 회로 대체 분석 데이터가 없습니다. 다시 분석을 실행해주세요." if lang_key == "ko"
                else "ℹ️ No interchangeability data. Please re-run the analysis.")

    # 종합 의견
    if result.get("recommendation"):
        st.subheader(t["recommendation"])
        st.markdown(f'<div class="recommend-box">{result["recommendation"]}</div>', unsafe_allow_html=True)

    if show_raw:
        with st.expander(t["json_raw"]): st.json(result)

    # 다운로드
    st.divider()
    st.subheader(t["download"])
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dl1, dl2 = st.columns(2)
    with dl1:
        try:
            excel_bytes = build_excel_report(result, lang_key)
            st.download_button(
                label=t["dl_excel"], data=excel_bytes,
                file_name=f"datasheet_report_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary",
            )
        except Exception as e:
            st.error(f"Excel 생성 오류: {e}")
    with dl2:
        st.download_button(
            label=t["dl_md"], data=build_markdown_report(result, lang_key).encode("utf-8"),
            file_name=f"datasheet_report_{timestamp}.md",
            mime="text/markdown", use_container_width=True,
        )
    st.caption(t["download_warning"])

render_result(st.session_state.result, t, lang_key, show_raw)

